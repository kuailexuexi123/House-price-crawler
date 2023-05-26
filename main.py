import pylab
from sklearn.preprocessing import LabelEncoder
from sklearn.preprocessing import LabelBinarizer
from sklearn.model_selection import train_test_split
import get_html_text as get
import numpy as np
import writefile
import writefile as wr
from bs4 import BeautifulSoup
import re
import csv
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import parsel
from pyecharts.charts import Map,Bar,Pie
from pyecharts import options as opts
import seaborn as sns
from sklearn.pipeline import FeatureUnion
from sklearn.preprocessing import StandardScaler
from pyecharts.globals import CurrentConfig, NotebookType
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.base import BaseEstimator, TransformerMixin
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import OneHotEncoder
p=[] # 总价格/万
hi=[]   # 有户型，平方数，朝向，款式，楼层，后面会进行string分割
fi=[]   # 关注度
jj=[]   # 每平米单价
name=[] # 房屋名称后续可以用来定位
qq1=[]
qq=[]
page=list(range(1,102,1))
url="https://quanzhou.lianjia.com/ershoufang/pg"
for i in page:
    response_text=get.get_html_text(url+str(i))
    wr.writefile("fj.txt",response_text)
    soup=BeautifulSoup(response_text,"html.parser")
    # 进行总价格的爬虫
    prices=soup.find_all('div',class_='priceInfo')
    for price in prices:
        p.append(price.span.string.strip())
    # 进行房屋整体信息的爬虫，外部为list，内部为string，需进行string分割获取每个的字段
    hs = soup.find_all('div', class_='houseInfo')
    for h in hs:
        hi.append(h.get_text()) # 在进行文本提取时不能使用.string的方法因为在这其中标签里覆盖标签。.string的方法无法得知是哪个标签
    # 关注度进行爬取
    followInfo=soup.find_all('div',class_='followInfo')
    for f in followInfo:
        fi.append(f.get_text()) # 与上面同
    # 价格进行爬取
    jj2=soup.find_all('div' ,class_="unitPrice" )
    for jj1 in jj2:
        jj.append(jj1.span.string.strip())
    # 爬取名字便于后续操作
    name2=soup.find_all('div',class_="positionInfo")
    for name1 in name2:
        for m in name1.stripped_strings:
            # print(m)
            qq.append(m)
# 接下来进入数据的处理，从爬取的数据中取出自己想要的
q=len(p)
fangchanxingming_list=[]
p_list=[]
pd_list=[]
guige_list=[]
daixao_list=[]
chaoxiang_list=[]
zhuangxiu_list=[]
louceng_list=[]
loucengyangshi_list=[]
guanzhudu_list=[]
diqu_list=[]
data=[]
sj=[]
e=0
m=0
for aa in qq:
    if m%3==0:
        name.append(aa)
    elif m%3==2:
        diqu_list.append(aa)
    m=m+1
for i in hi:
    c=i.split("|")
    guige_list.append(c[0].strip())
    d=re.findall(r'\d+', c[1])
    daixao_list.append(d[0].strip())
    a="".join(c[2].split())
    chaoxiang_list.append(a)
    zhuangxiu_list.append(c[3].strip())
    louceng_list.append(c[4].strip())
    loucengyangshi_list.append(c[5].strip())
for i in p:
    p_list.append(i)
for j in jj:
    j1=re.findall(r'\d+', j)
    j2=j1[0]+j1[1]
    # j2=re.findall(r'\d', j)
    pd_list.append(j2)
for l in fi:
    ll=l.split("/")
    lll=ll[0]
    lll1=re.findall(r'\d+',lll) # 引入正则表达式搜索字符串
    guanzhudu_list.append(int(lll1[0])*20+30)
for x in name:
     fangchanxingming_list.append(x)
wb2=openpyxl.load_workbook('map-location.xlsx')
st1=wb2["位置"]
jd=[]
wd=[]
for l in st1['B']:
    jd.append(l.value)
for l1 in st1['C']:
    wd.append(l1.value)
for k in range(0,q):
    data.append([fangchanxingming_list[k],diqu_list[k],jd[k],wd[k],p_list[k],pd_list[k],guige_list[k],daixao_list[k],chaoxiang_list[k],zhuangxiu_list[k],louceng_list[k],loucengyangshi_list[k],guanzhudu_list[k]])


# 进行处理后，将处理后的数据写入csv和excel中，后续通过pandas对csv文件进行数据帧的读取
filename="fjjg.csv"
with open(filename, 'w', newline='', encoding='utf-8') as f:  # 设置newline,否则两行之间会有空一行
    writer = csv.writer(f)
    writer.writerow(["名称",'地区','经度','维度',"总价/万","单价/元","规格","大小/平方米","朝向","装修","楼层","楼层样式","关注人数"])
    for row in data:
        writer.writerow(row)
with open("fjjg.csv",'r',encoding='utf-8')as f:
    reader=csv.reader(f)
    for r in reader:
        if reader.line_num==1:
            continue
        else:
            out = '{\"lng\":'+r[2]+',\"lat\":'+r[3]+',\"count\":'+r[12]+'},'##热力图数据的采集与拼凑
with open("fjjg.csv",'r',encoding='utf-8')as f:
    reader=csv.reader(f)
    for r in reader:
        if reader.line_num==1:
            continue
        else:
            out1 = '{\"lng\":'+r[2]+',\"lat\":'+r[3]+',\"count\":'+r[5]+'},'##热力图数据的采集与拼凑
wb = openpyxl.Workbook()# 新建一个Excel工作簿
wb.create_sheet('房产价格',0)# 在前面新建一个工作表
wb.save('fjpq10.xlsx')# 为工作簿命名
wb1 = openpyxl.load_workbook('fjpq10.xlsx')
print(wb1.read_only,wb1.active)
st = wb["房产价格"]
one_row=["名称","地区",'经度','维度',"总价/万","单价/元","规格","大小/平方米","朝向","装修","楼层","楼层样式","关注人数"]
st.append(one_row)
for i in range(0,q):
    sj=[fangchanxingming_list[i],diqu_list[i],jd[i],wd[i],p_list[i],pd_list[i],guige_list[i],daixao_list[i],chaoxiang_list[i],zhuangxiu_list[i],louceng_list[i],loucengyangshi_list[i],guanzhudu_list[i]]
    st.append(sj)
wb.save('fjpq1.xlsx')# 为工作簿命名



# #绘制兰丁格尔图
df = pd.read_csv('fjjg.csv',sep=',',encoding='utf-8')
df.head()
df_counts = df.groupby('地区')['总价/万'].count() # 对地区进行分组，统计该地区有多少房子在售出，此处的总价/万字段只是用来统计数目选用其他的也行
df0 = df_counts.copy()
df0.sort_values(ascending=False, inplace=True)#对其进行排序
name = df_counts.index.tolist()
count = df_counts.values.tolist()
c0 = (
    Pie()
    .add(
        '',
        [list(z) for z in zip(name, count)],
        # 饼图的半径，数组的第一项是内半径，第二项是外半径
        # 默认设置成百分比，相对于容器高宽中较小的一项的一半
        radius=['20%', '60%'],
        # 让图在这个位置显示
        center=['50%', '65%'],
        # 是否展示成南丁格尔图，通过半径区分数据大小，有'radius'和'area'两种模式。
        # radius：扇区圆心角展现数据的百分比，半径展现数据的大小
        # area：所有扇区圆心角相同，仅通过半径展现数据大小
        rosetype="radius",
        # 显示标签
        label_opts=opts.LabelOpts(is_show=False),
    )
    .set_series_opts(label_opts=opts.LabelOpts(formatter='{b}: {c}'))
)
c0.render()


# 计算每个地区的平均房价
df_mean=df.groupby('地区')['单价/元'].mean() #利用pandas算出分组每个地区房价的平均图并进行绘制
df1=df_mean.copy()
diqu_list=df1.index.tolist() #将索引列转换成列表
danjia_list=df1.values.tolist()# 将索引列转换成列表
danjia1_list=[]
for i in danjia_list:
    i=int(i)
    danjia1_list.append(i)
df_max=df.groupby('地区')['单价/元'].max()#利用pandas算出每个分组的最大值
danjiamax_list=df_max.values.tolist()



#绘制房价柱状图
# 解决中文显示问题 指定默认字体
plt.rcParams['font.sans-serif'] = ['KaiTi']
l1=danjia1_list
# print(l1)
# print(len(l1))
l2=danjiamax_list
name=diqu_list
x=list(range(0,len(name)))
total_width, n = 0.8, 2
width = total_width / n
a=plt.bar(x, l1, width=width, label='平均单价/元',fc = 'y')
for i in range(len(x)):
    x[i] = x[i] + width
b=plt.bar(x, l2, width=width, label='最高单价/元',tick_label = name,fc = 'r')
plt.xticks(rotation=45)
plt.xlabel('泉州地区')
plt.ylabel('房子价格/元')
plt.title('各个地区房子均价与最高价')
plt.legend()
plt.show()


#绘制折线图
df3=df[df["单价/元"]<=10000]
aa=df3["名称"].count()
bb=df[(df["单价/元"]<=20000) & (df["单价/元"]>10000)]["名称"].count()
cc=b=df[(df["单价/元"]<=30000) & (df["单价/元"]>20000)]["名称"].count()
dd=b=df[(df["单价/元"]<=40000) & (df["单价/元"]>30000)]["名称"].count()
ee=b=df[(df["单价/元"]<=50000) & (df["单价/元"]>40000)]["名称"].count()
y1=[aa,bb,cc,dd,ee]
x1=range(10000,60000,10000)
plt.plot(x1,y1,label='各个区间里的个数',linewidth=3,color='r',marker='o',
markerfacecolor='blue',markersize=12)
plt.xlabel('单价/元（区间内）')
plt.ylabel('数量')
plt.title('各个区间里的房屋个数')
for a, b in zip(x1,y1):
    plt.text(a,b,b,ha='center',va='bottom',fontsize=10)
plt.legend()
plt.show()


#下面为对房价进行机器学习预测部分
#先对数据处理，处理后判断其相关性
def split_train_test(data, test_ratio):
    np.random.seed(42)
    shuffled_indices = np.random.permutation(len(data))
    test_set_size = int(len(data) * test_ratio)
    test_indices = shuffled_indices[:test_set_size]
    train_indices = shuffled_indices[test_set_size:]
    return data.iloc[train_indices], data.iloc[test_indices]
train_set, test_set = split_train_test(df, 0.2)  ##该方法为pandas进行训练集与测试集的划分
train_set, test_set = train_test_split(df, test_size=0.2, random_state=42) #该方法为skl里的划分方法
housing_num = df.drop(['名称','地区','规格','朝向','装修','楼层','楼层样式'], axis=1) #取出数值部分
housing_num['总价/元']= np.ceil(housing_num['总价/万'] *10000)#进行线性回归时万的单位太小，换成统一元
housing_num=housing_num.drop("总价/万",axis=1)
imputer = SimpleImputer(strategy='median')#进行中位数计算，因为观看中位数的数据更有代表性
imputer.fit(housing_num)#对数据进行中位数观看
imputer.statistics_
print(housing_num.median().values)#对数据的中位数进行展示
X = imputer.transform(housing_num)
housing_tr = pd.DataFrame(X, columns=housing_num.columns)
print(housing_tr.describe())#对数据进行观看，平均值，1/4，1/2，3/4的值
df.hist(bins=50,figsize=(20,15))#对数值类型进行直方图绘图
plt.show()
corr_matrix = housing_num.corr()
print(corr_matrix["总价/元"].sort_values(ascending=False))#总价/元
x_simple = np.array(housing_num['总价/元'])
y_simple = np.array(housing_num['大小/平方米'])
my_rho = np.corrcoef(x_simple, y_simple)#相关性计算，先研究变量与目标的相关性，有的相关性低的可以剔除，用numpy利用了皮尔逊公式
z_simple = np.array(housing_num['单价/元'])
my_rho1 = np.corrcoef(x_simple, z_simple)
print(my_rho1)
print(my_rho)
print(housing_num['总价/元'].describe()) #对目标变量进行
sns.displot(housing_num['单价/元'])
sns.jointplot(x='大小/平方米',y='总价/万',data=df)
sns.jointplot(x='单价/元',y='总价/万',data=df)
df.info()#判断数据有无丢失，结果为基本无数据缺失
#下面为画出各个数据的相关性
corrmat=df.corr()
fig,ax=plt.subplots(figsize=(12,10))
sns.heatmap(corrmat,square=True)#画出相关性热点图

plt.show()
#对分类型变量需进行编码，如房屋结构之类的
encoder = LabelEncoder()
encoder1 = LabelEncoder()
housing_cat = df['朝向']
housing_cat_encoded = encoder.fit_transform(housing_cat)
housing_cat1 = df['楼层样式']
housing_cat2 = df['装修']
housing_cat3 = df['规格']
housing_cat_encoded1 = encoder1.fit_transform(housing_cat1)
# for i in housing_cat_encoded:
#     print(i)
print(housing_cat_encoded)#
print(encoder.classes_)
print(housing_cat_encoded1)
print(encoder1.classes_)
#上面对类别进行编码，因为类别与数值大小无关，所以需要转变为独热编码
encoder2 = LabelBinarizer()
housing_cat_1hot2 = encoder2.fit_transform(housing_cat1)
housing_cat_1hot = encoder2.fit_transform(housing_cat2)
housing_cat_1hot3 = encoder2.fit_transform(housing_cat3)
print(housing_cat_1hot2)
print(housing_cat_1hot)
print(housing_cat_1hot3)
#数据处理完毕后，接下来需要进行特征放缩和数据融合拼凑进行线性回归
#接下来借鉴了转换器，可以进行数值类型和分类类型的数据进行统一转换
# 选取列名
col_names = ["经度", "维度", "大小/平方米",'关注人数']
rooms_ix, bedrooms_ix, population_ix, households_ix = [df.columns.get_loc(c) for c in col_names]


class CombinedAttributesAdder(BaseEstimator, TransformerMixin):
    def __init__(self, add_bedrooms_per_room=True):  # no *args or **kargs
        self.add_bedrooms_per_room = add_bedrooms_per_room

    def fit(self, X, y=None):
        return self

    def transform(self, X):
        rooms_per_household = X[:, rooms_ix] / X[:, households_ix]
        population_per_household = X[:, population_ix] / X[:, households_ix]
        # 根据超参数add_bedrooms_per_room判断是否需要添加该组合属性
        if self.add_bedrooms_per_room:
            bedrooms_per_room = X[:, bedrooms_ix] / X[:, rooms_ix]
            return np.c_[X, rooms_per_household, population_per_household,
                         bedrooms_per_room]
        else:
            return np.c_[X, rooms_per_household, population_per_household]


attr_adder = CombinedAttributesAdder(add_bedrooms_per_room=False)
housing_extra_attribs = attr_adder.transform(df.values)
# 将housing_extra_attribs从array转为DataFrame
housing_extra_attribs = pd.DataFrame(
    housing_extra_attribs,
    columns=list(df.columns) + ["rooms_per_household", "population_per_household"],
    index=df.index)
housing_extra_attribs.head()
num_pipeline = Pipeline([
		# 中位数替换缺失值
        ('imputer', SimpleImputer(strategy="median")),
        # 添加组合属性
        ('attribs_adder', CombinedAttributesAdder()),
        # 归一化，统一量纲
        ('std_scaler', StandardScaler()),
    ])

housing_num_tr = num_pipeline.fit_transform(housing_num)
housing_num_tr
from sklearn.compose import ColumnTransformer

# 获得数值列名称列表
num_attribs = list(housing_num)
# 获得类别列名称列表
cat_attribs = ["规格","朝向","装修","楼层","楼层样式"]

# 元组中的三个参数分别代表：名称（自定），转换器，以及一个该转换器能够应用的列名字（或索引）的列表
full_pipeline = ColumnTransformer([
		# 数值属性列转换器
        ("num", num_pipeline, num_attribs),
        # 文本属性列转换器
        ("cat", OneHotEncoder(), cat_attribs),
    ])
# 将ColumnTranformer应用到房屋数据
housing_prepared = full_pipeline.fit_transform(df)
print(housing_prepared)
##太多分数值类。。。。



